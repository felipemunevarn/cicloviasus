from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A4'] = 1
ws['D1'] = 4
wb.save('test.xlsx')