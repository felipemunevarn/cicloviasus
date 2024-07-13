from openpyxl import Workbook
from openpyxl import cell
from catalogo.models import Producto

def create_excel(request, daily_cart, customer):
    if (request != ""):
        cart = request.session.get("carro")
    else:
        cart = daily_cart
    user = request.user
    delivery_date = request.POST.get("deliveryDate")
    comments = request.POST.get("comments")
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Codigo'
    ws['B1'] = 'Nombre'
    ws['C1'] = 'Cantidad'
    ws['D1'] = 'Vendedor'
    ws['E1'] = 'Cliente'
    ws['F1'] = 'FechaEntrega'
    ws['G1'] = 'Comentarios'
    row = 2
    column = 1
    column_widths = [6,6,7,8,7,12,11]
    for item in cart:
        producto = Producto.objects.get(pk=item)
        ws.cell(row=row, column=column, value=producto.codigo)
        if (len(producto.codigo) > column_widths[column-1]):
            column_widths[column-1] = len(producto.codigo)
        column += 1 
        ws.cell(row=row, column=column, value=producto.titulo)
        if (len(producto.titulo) > column_widths[column-1]):
            column_widths[column-1] = len(producto.titulo)
        column += 1 
        ws.cell(row=row, column=column, value=cart[item]["cantidad"])
        column += 1 
        ws.cell(row=row, column=column, value=user.username)
        if (len(user.username) > column_widths[column-1]):
            column_widths[column-1] = len(user.username)
        column += 1 
        ws.cell(row=row, column=column, value=customer.nombre)
        if (len(customer.nombre) > column_widths[column-1]):
            column_widths[column-1] = len(customer.nombre)
        column += 1 
        ws.cell(row=row, column=column, value=delivery_date)
        if (len(delivery_date) > column_widths[column-1]):
            column_widths[column-1] = len(delivery_date)
        column += 1 
        ws.cell(row=row, column=column, value=comments)
        if (len(comments) > column_widths[column-1]):
            column_widths[column-1] = len(comments)
        column = 1 
        row += 1

    # Calculate column widths based on cell content

    for i, column_width in enumerate(column_widths, 1):
        column_letter = chr(64 + i)  # Convert index to column letter
        ws.column_dimensions[column_letter].width = column_width + 5
    wb.save('./report.xlsx')
    wb.close()

def resume_excel(cart):
    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'Codigo'
    ws['B1'] = 'Titulo'
    ws['C1'] = 'Cantidad'

    row = 2
    column = 1

    column_widths = [6,6,8]

    for prod,atts in cart.items():

        ws.cell(row=row, column=column, value=atts['code'])
        if (len(atts['code']) > column_widths[column-1]):
            column_widths[column-1] = len(atts['code'])
        column += 1 

        ws.cell(row=row, column=column, value=atts['title'])
        if (len(atts['title']) > column_widths[column-1]):
            column_widths[column-1] = len(atts['title'])
        column += 1

        ws.cell(row=row, column=column, value=atts['qty'])
        
        column = 1 
        row += 1

    # Calculate column widths based on cell content

    for i, column_width in enumerate(column_widths, 1):
        column_letter = chr(64 + i)  # Convert index to column letter
        ws.column_dimensions[column_letter].width = column_width + 5
    wb.save('./report.xlsx')
    wb.close()
